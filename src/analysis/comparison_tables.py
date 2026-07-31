"""
Generate comparison tables for experiments.

Outputs

1. CSV files
2. Excel workbook
3. Pandas DataFrames
"""

from pathlib import Path

import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

class ComparisonTables:

    def __init__(self, analyzer):

        self.analyzer = analyzer

        self.output_dir = analyzer.output_dir

    def operational_dataframe(self):

        metrics = self.analyzer.compare_metrics()

        df = pd.DataFrame(metrics).T

        df = df.reset_index()

        df.rename(
            columns={
                "index": "Metric"
            },
            inplace=True
        )

        return df

    def action_distribution_dataframe(self):

        data = self.analyzer.compare_actions()

        df = pd.DataFrame(data)

        df.index.name = "Action"

        return df.reset_index()

    def successful_actions_dataframe(self):

        data = self.analyzer.compare_successful_actions()

        df = pd.DataFrame(data)

        df.index.name = "Successful Action"

        return df.reset_index()

    def incident_dataframe(self):

        data = self.analyzer.compare_incidents()

        df = pd.DataFrame(data)

        df.index.name = "Incident"

        return df.reset_index()

    def success_by_incident_dataframe(self):

        rows = {}

        structured = self.analyzer.structured.get(
            "success_by_incident",
            {}
        )

        fixed = self.analyzer.fixed.get(
            "success_by_incident",
            {}
        )

        adaptive = self.analyzer.adaptive.get(
            "success_by_incident",
            {}
        )

        incidents = sorted(
            set(structured.keys())
            |
            set(fixed.keys())
            |
            set(adaptive.keys())
        )

        for incident in incidents:

            rows[incident] = {

                "Structured":
                    structured.get(incident),

                "Fixed Alpha":
                    fixed.get(incident),

                "Adaptive Alpha":
                    adaptive.get(incident)

            }

        df = pd.DataFrame(rows).T

        df.index.name = "Incident"

        return df.reset_index()

    def performance_dataframe(self):

        performance = self.analyzer.adaptive_performance()

        rows = []

        for metric, values in performance.items():

            if isinstance(values, dict):

                rows.append({

                    "Metric": metric,

                    "Average":
                        values.get("average"),

                    "Minimum":
                        values.get("minimum"),

                    "Maximum":
                        values.get("maximum"),

                    "Samples":
                        values.get("count")

                })

        return pd.DataFrame(rows)

    def improvement_dataframe(self):
        """
        Compare the percentage improvement between
        Structured, Fixed Alpha and Adaptive Alpha.

        Positive values indicate an increase.

        For MTTR, a negative value is desirable because
        lower MTTR means faster incident resolution.
        """

        metrics = self.analyzer.compare_metrics()

        rows = []

        for metric, values in metrics.items():

            structured = values["Structured"]
            fixed = values["Fixed Alpha"]
            adaptive = values["Adaptive Alpha"]

            fixed_vs_structured = (
                (fixed - structured) / structured
            ) * 100

            adaptive_vs_structured = (
                (adaptive - structured) / structured
            ) * 100

            adaptive_vs_fixed = (
                (adaptive - fixed) / fixed
            ) * 100

            rows.append({

                "Metric": metric,

                "Fixed vs Structured (%)":
                    round(fixed_vs_structured, 2),

                "Adaptive vs Structured (%)":
                    round(adaptive_vs_structured, 2),

                "Adaptive vs Fixed (%)":
                    round(adaptive_vs_fixed, 2)

            })

        return pd.DataFrame(rows)

    def export_csv(self):

        self.operational_dataframe().to_csv(

            self.output_dir /
            "comparison_summary.csv",

            index=False
        )

        self.action_distribution_dataframe().to_csv(

            self.output_dir /
            "action_distribution.csv",

            index=False
        )

        self.successful_actions_dataframe().to_csv(

            self.output_dir /
            "successful_actions.csv",

            index=False
        )

        self.incident_dataframe().to_csv(

            self.output_dir /
            "incident_distribution.csv",

            index=False
        )

        self.success_by_incident_dataframe().to_csv(

            self.output_dir /
            "success_by_incident.csv",

            index=False
        )

        self.performance_dataframe().to_csv(

            self.output_dir /
            "adaptive_performance.csv",

            index=False
        )

        self.improvement_dataframe().to_csv(
            self.output_dir /
            "improvement_summary.csv",
            index=False
        )

    def export_excel(self):

        file = self.output_dir / "comparison_summary.xlsx"

        with pd.ExcelWriter(file, engine="openpyxl") as writer:

            self.operational_dataframe().to_excel(

                writer,

                sheet_name="Operational",

                index=False
            )

            self.action_distribution_dataframe().to_excel(

                writer,

                sheet_name="Actions",

                index=False
            )

            self.successful_actions_dataframe().to_excel(

                writer,

                sheet_name="Successful Actions",

                index=False
            )

            self.incident_dataframe().to_excel(

                writer,

                sheet_name="Incidents",

                index=False
            )

            self.success_by_incident_dataframe().to_excel(

                writer,

                sheet_name="Incident Success",

                index=False
            )

            self.performance_dataframe().to_excel(

                writer,

                sheet_name="Performance",

                index=False
            )
            self.improvement_dataframe().to_excel(
                writer,
                sheet_name="Improvement",
                index=False
            )

            workbook = writer.book

            for sheet in workbook.sheetnames:

                ws = workbook[sheet]

                for cell in ws[1]:

                    cell.font = Font(bold=True)

                for column in ws.columns:

                    length = max(
                        len(str(cell.value))
                        if cell.value is not None else 0
                        for cell in column
                    )

                    ws.column_dimensions[
                        get_column_letter(column[0].column)
                    ].width = min(length + 3, 40)

    def generate(self):

        print()

        print("=" * 60)

        print("Generating comparison tables...")

        self.export_csv()

        self.export_excel()

        print("Completed.")

        print("=" * 60)